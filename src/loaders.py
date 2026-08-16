"""
Belge okuyucular (PDF / DOCX / TXT / MD / CSV / XLSX).

TASARIM İLKESİ: Her metin bloğu, kaynağına geri izlenebilir olmalıdır.
Bu yüzden her blok şu metadata ile döner:
    source_file : dosya adı
    source_path : göreli yol
    page        : PDF sayfa numarası (1'den başlar) | DOCX'te None
    locator     : insan tarafından okunabilir konum ("Sayfa 12", "Paragraf 45", "Tablo 3")
    order       : belge içi sıra numarası (chunk birleştirme için)
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterator, List, Optional

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".txt", ".md",
                        ".csv", ".tsv", ".xlsx", ".xlsm", ".xls"}


@dataclass
class Block:
    """Belgeden çıkarılmış, konumu bilinen ham metin bloğu."""
    text: str
    page: Optional[int] = None
    locator: str = ""
    order: int = 0
    extra: Dict[str, str] = field(default_factory=dict)


# ---------------------------------------------------------------- yardımcılar

_WS_RE = re.compile(r"[ \t ]+")
_MULTI_NL_RE = re.compile(r"\n{3,}")
# PDF'lerde satır sonu tirelemesi: "kon-\nsantre" -> "konsantre"
_HYPHEN_RE = re.compile(r"(\w)-\n(\w)")
# Sayfa altı/üstü tekrarlayan gürültü (sayfa numarası tek başına satırda)
_PAGENUM_RE = re.compile(r"^\s*[-–—]?\s*\d{1,4}\s*[-–—]?\s*$", re.MULTILINE)


def clean_text(raw: str) -> str:
    if not raw:
        return ""
    t = raw.replace("\r\n", "\n").replace("\r", "\n")
    t = _HYPHEN_RE.sub(r"\1\2", t)
    t = _PAGENUM_RE.sub("", t)
    t = _WS_RE.sub(" ", t)
    t = _MULTI_NL_RE.sub("\n\n", t)
    # Türkçe karakter bozulmalarını yakalamak için basit kontrol bırakıldı.
    return t.strip()


def detect_encoding(path: Path, default: str = "utf-8") -> str:
    try:
        import chardet  # type: ignore
        raw = path.read_bytes()[:65536]
        guess = chardet.detect(raw)
        return guess.get("encoding") or default
    except Exception:
        return default


# ---------------------------------------------------------------- okuyucular

# --------------------------------------------------------------------------
# PDF İÇİ TABLO SATIRI TESPİTİ
# --------------------------------------------------------------------------
# PDF'ler tabloları düz metin satırları olarak verir. Bir sayfayı tek parça
# olarak indekslersek, "2021'in 5. ayı" sorulduğunda model aynı bloktaki
# komşu satırı (6. ay) okuyup YANLIŞ tutarı verebilir. Bu yüzden tablo
# görünümlü satırları tespit edip her birini bağımsız parça yapıyoruz.

_TR_MONTHS = ("Ocak|Şubat|Mart|Nisan|Mayıs|Haziran|Temmuz|Ağustos|Eylül|Ekim|Kasım|Aralık")

# Türkçe biçimli sayı: 1.490,00 · 1490,50 · 12 · %5
_NUM_RE = re.compile(r"(?<![\w.,])-?\d{1,3}(?:\.\d{3})+(?:,\d+)?(?![\w])"
                     r"|(?<![\w.,])-?\d+(?:,\d+)?(?![\w])")

# Tarih / dönem belirteci: 01.05.2021 · 2021-05 · 05/2021 · Mayıs 2021 · 2021 Mayıs
_PERIOD_RE = re.compile(
    r"\b(?:\d{1,2}[./-]\d{1,2}[./-]\d{2,4}"
    rf"|\d{{4}}[-/.]\d{{1,2}}\b"
    rf"|\b\d{{1,2}}[-/.]\d{{4}}"
    rf"|(?:{_TR_MONTHS})\s*[/-]?\s*\d{{4}}"
    rf"|\d{{4}}\s+(?:{_TR_MONTHS}))\b",
    re.IGNORECASE,
)


def _is_table_row(line: str) -> bool:
    """Satır bir tablo verisi mi? (başlık/paragraf değil)"""
    s = line.strip()
    if not (4 <= len(s) <= 220):
        return False
    nums = _NUM_RE.findall(s)
    has_period = bool(_PERIOD_RE.search(s))
    if has_period and len(nums) >= 1:
        return True
    if len(nums) >= 3:
        return True
    # "Ad ... 1.234,00" gibi iki sütunlu satırlar: az kelime + para biçimli sayı
    if len(nums) >= 1 and len(s.split()) <= 10 and re.search(r"\d,\d{2}\b", s):
        return True
    return False


def _split_page_rows(text: str, page_no: int, order_base: int) -> List[Block]:
    """
    Sayfayı tablo satırları + anlatı metni olarak ikiye ayırır.
    Tablo satırı sayısı 3'ten azsa sayfa bölünmez (yanlış pozitifi önler).
    """
    lines = [l for l in text.split("\n") if l.strip()]
    row_idx = [i for i, l in enumerate(lines) if _is_table_row(l)]

    if len(row_idx) < 3:
        return []

    # Başlık satırı: ilk veri satırından hemen önceki, veri olmayan satır
    header = ""
    first = row_idx[0]
    for i in range(first - 1, max(-1, first - 4), -1):
        cand = lines[i].strip()
        if cand and not _is_table_row(cand) and len(cand) <= 160:
            header = cand
            break

    blocks: List[Block] = []
    order = order_base

    # 1) Anlatı metni (tablo dışı satırlar) tek blok
    narrative = "\n".join(l for i, l in enumerate(lines) if i not in set(row_idx)).strip()
    if len(narrative) >= 60:
        order += 1
        blocks.append(Block(text=narrative, page=page_no,
                            locator=f"Sayfa {page_no}", order=order))

    # 2) Her tablo satırı bağımsız blok
    for n, i in enumerate(row_idx, start=1):
        order += 1
        body = lines[i].strip()
        text_out = (f"[TABLO SATIRI] {header}\n{body}" if header
                    else f"[TABLO SATIRI] {body}")
        blocks.append(Block(
            text=text_out,
            page=page_no,
            locator=f"Sayfa {page_no}, satır {n}",
            order=order,
            extra={"type": "table"},
        ))
    return blocks


# --------------------------------------------------------------------------
# ÜSTBİLGİ / ALTBİLGİ (BOILERPLATE) AYIKLAMA
# --------------------------------------------------------------------------
# Kurumsal belgelerin hemen her sayfasında aynı üstbilgi/altbilgi bulunur
# ("ADL-2024/117 — Merkez Depo Hizmet Sözleşmesi    Sayfa 7"). Bu satırlar
# içerik değildir ama indekse girer ve iki ayrı zarar verir:
#
#   1. Tablo satırı bölme açıkken her sayfa için AYRI BİR PARÇA oluşur.
#      20 sayfalık bir belgede 20 çöp parça demektir.
#   2. Bu parçalar sözleşme numarası ve belge adı içerdiği için, o terimleri
#      taşıyan sorularda BM25 skoru alır ve ilk sıralara girer. Gerçek bir
#      ölçümde tam olarak bu oldu: dört kaynak slotundan biri altbilgi
#      parçasına gitti ve doğru satırın yerini kaptı.
#
# Sabit kalıp yazmak yerine TEKRAR tespit edilir: sayfaların çoğunda, hem de
# sayfa başında/sonunda görünen satır boilerplate'tir. Sayfa numaraları
# değiştiği için karşılaştırmadan önce rakamlar maskelenir.

_BP_HEAD_TAIL = 3          # sayfa başı/sonunda kaç satıra bakılacak
_BP_MIN_LEN = 12           # daha kısa satırlar zaten anlamsız
_BP_MIN_PAGES = 3          # bu kadar az sayfada tespit güvenilir değil


def _norm_line(line: str) -> str:
    """Karşılaştırma için satırı normalleştirir (sayfa numarası maskelenir)."""
    s = re.sub(r"\d+", "#", line)
    return re.sub(r"\s+", " ", s).strip().lower()


def _detect_boilerplate(page_texts: List[str], min_ratio: float = 0.6) -> set:
    """Sayfaların çoğunda başta/sonda tekrarlayan satırları bulur."""
    if len(page_texts) < _BP_MIN_PAGES:
        return set()
    from collections import Counter
    sayac: Counter = Counter()
    for t in page_texts:
        satirlar = [l for l in t.splitlines() if l.strip()]
        aday = satirlar[:_BP_HEAD_TAIL] + satirlar[-_BP_HEAD_TAIL:]
        for norm in {_norm_line(l) for l in aday}:
            if len(norm) >= _BP_MIN_LEN:
                sayac[norm] += 1
    esik = max(_BP_MIN_PAGES, int(len(page_texts) * min_ratio))
    return {l for l, n in sayac.items() if n >= esik}


def _strip_boilerplate(text: str, bp: set) -> str:
    """Tespit edilen boilerplate satırlarını YALNIZCA sayfa başı/sonundan atar.

    Konum kısıtı önemlidir: aynı ifade sayfa ortasında gerçek içerik olarak
    geçebilir (ör. sözleşme numarasının metin içinde anılması). Ortadaki
    kullanımlar korunur.
    """
    if not bp:
        return text
    satirlar = text.splitlines()
    n = len(satirlar)
    tut: List[str] = []
    for i, l in enumerate(satirlar):
        bas_ya_da_son = i < _BP_HEAD_TAIL or i >= n - _BP_HEAD_TAIL
        if bas_ya_da_son and l.strip() and _norm_line(l) in bp:
            continue
        tut.append(l)
    return "\n".join(tut).strip()


def load_pdf(path: Path,
             split_table_rows: bool = True,
             ocr_options: Optional[Dict[str, object]] = None,
             progress=None,
             warnings: Optional[List[Dict[str, object]]] = None) -> List[Block]:
    """
    pypdf ile sayfa sayfa okur. Sayfa numarası atıf için kritiktir.

    Metin katmanı olmayan (taranmış) sayfalar tespit edilip, OCR açıksa
    görüntüden metne çevrilir. Karma belgelerde sayfa sayfa karar verilir.
    """
    from pypdf import PdfReader

    opts = ocr_options or {}
    ocr_enabled = bool(opts.get("enabled", True))
    min_chars = int(opts.get("min_chars_per_page", 60))
    lang = str(opts.get("language", "tur+eng"))
    dpi = int(opts.get("dpi", 300))
    preprocess = bool(opts.get("preprocess", True))
    preserve_spaces = bool(opts.get("preserve_spaces", True))

    blocks: List[Block] = []
    reader = PdfReader(str(path))

    if reader.is_encrypted:
        try:
            reader.decrypt("")  # boş parolalı koruma
        except Exception as exc:
            raise RuntimeError(f"Şifreli PDF açılamadı: {path.name} ({exc})")

    total = len(reader.pages)
    order = 0
    ocr_used = 0
    ocr_failed_reason = ""
    scanned_pages = 0
    low_quality: List[Dict[str, object]] = []

    # 1. GEÇİŞ — tüm sayfaların metnini çıkar (gerekirse OCR ile).
    # Boilerplate tespiti sayfaları KARŞILAŞTIRMAYA dayandığı için, parçalama
    # yapılmadan önce bütün sayfaların elde olması gerekir.
    sayfa_metinleri: List[str] = []
    sayfa_ocr: List[bool] = []

    for idx, page in enumerate(reader.pages, start=1):
        text = ""
        bu_sayfa_ocr = False
        # DÜZEN KORUMALI ÇIKARIM: varsayılan mod tablo hücrelerini ayrı
        # satırlara dağıtıyor ve bir tablo satırı ("11/2024 | Envanter
        # sayım | 36 | 1.455.200,00") parça parça kopuyor; sonuçta indekse
        # yalnızca "11/2024" giriyor, tutar kayboluyor. "layout" modu
        # hücreleri aynı satırda tutar. Eski pypdf sürümlerinde bu kip
        # yoksa sessizce klasik moda düşülür.
        try:
            text = clean_text(page.extract_text(extraction_mode="layout") or "")
        except Exception:
            text = ""
        if len(text) < 40:
            try:
                text = clean_text(page.extract_text() or "")
            except Exception:
                text = ""

        # --- Taranmış sayfa mı? ---
        if len(text) < min_chars:
            scanned_pages += 1
            if ocr_enabled:
                from . import ocr as ocr_mod
                ok, reason = ocr_mod.availability(opts.get("tesseract_cmd") or None)
                if ok:
                    if progress:
                        progress(f"{path.name}: sayfa {idx}/{total} OCR ile okunuyor…")
                    ocr_text = ocr_mod.ocr_page(
                        path, idx - 1, lang=lang, dpi=dpi,
                        preprocess=preprocess, preserve_spaces=preserve_spaces)
                    if len(ocr_text) > len(text):
                        text = clean_text(ocr_text)
                        ocr_used += 1
                        bu_sayfa_ocr = True
                        # Bozuk OCR sessizce indekse girip yanlış sayı
                        # üretilmesine yol açıyor. Şüpheli sayfaları
                        # kullanıcıya bildirebilmek için işaretle.
                        score, problems = ocr_mod.assess_quality(text)
                        if score < float(opts.get("min_quality", 0.65)):
                            low_quality.append({
                                "page": idx,
                                "score": round(score, 2),
                                "issues": problems,
                            })
                else:
                    ocr_failed_reason = reason

        sayfa_metinleri.append(text)
        sayfa_ocr.append(bu_sayfa_ocr)

    # 2. GEÇİŞ — tekrarlayan üstbilgi/altbilgi satırlarını ayıkla, sonra parçala.
    boilerplate = _detect_boilerplate(sayfa_metinleri)

    for idx, (text, sayfa_ocrlu) in enumerate(zip(sayfa_metinleri, sayfa_ocr), start=1):
        text = _strip_boilerplate(text, boilerplate)
        if not text:
            continue

        # NOT: OCR işareti artık SAYFA BAZINDA. Önceden toplam sayaç
        # kullanılıyordu; ilk OCR'lı sayfadan sonraki tüm sayfalar —dijital
        # olsalar bile— "ocr" damgası alıyordu ve karma belgelerde hangi
        # sayfanın gerçekten taranmış olduğu kaybediliyordu.
        isaret = {"ocr": "1"} if sayfa_ocrlu else {}

        if split_table_rows:
            row_blocks = _split_page_rows(text, idx, order)
            if row_blocks:
                order = row_blocks[-1].order
                for b in row_blocks:
                    if sayfa_ocrlu:
                        b.extra.setdefault("ocr", "1")
                blocks.extend(row_blocks)
                continue

        order += 1
        blocks.append(Block(text=text, page=idx, locator=f"Sayfa {idx}",
                            order=order, extra=isaret))

    if warnings is not None and low_quality:
        warnings.extend(low_quality)

    if not blocks:
        if scanned_pages and ocr_failed_reason:
            raise RuntimeError(
                f"'{path.name}' taranmış (görüntü) bir PDF ve OCR yapılamadı. "
                f"{ocr_failed_reason}"
            )
        if scanned_pages:
            raise RuntimeError(
                f"'{path.name}' taranmış (görüntü) bir PDF. OCR devre dışı "
                "(config.yaml → ocr.enabled: true yapın) veya OCR metin üretemedi. "
                "Sayfa kalitesi düşükse ocr.dpi değerini 300'e çıkarmayı deneyin."
            )
        raise RuntimeError(f"'{path.name}' dosyasından metin çıkarılamadı.")

    return blocks


def load_docx(path: Path) -> List[Block]:
    """python-docx ile paragraf ve tabloları sırayla okur."""
    from docx import Document  # type: ignore
    from docx.table import Table  # type: ignore
    from docx.text.paragraph import Paragraph  # type: ignore

    doc = Document(str(path))
    blocks: List[Block] = []
    para_no = 0
    table_no = 0
    order = 0

    body = doc.element.body
    for child in body.iterchildren():
        tag = child.tag.split("}")[-1]

        if tag == "p":
            para_no += 1
            p = Paragraph(child, doc)
            text = clean_text(p.text)
            if not text:
                continue
            order += 1
            style = (p.style.name or "") if p.style is not None else ""
            is_heading = style.lower().startswith(("heading", "başlık"))
            blocks.append(Block(
                text=text,
                page=None,
                locator=f"Paragraf {para_no}",
                order=order,
                extra={"style": style, "heading": "1" if is_heading else "0"},
            ))

        elif tag == "tbl":
            table_no += 1
            tbl = Table(child, doc)
            rows: List[str] = []
            for row in tbl.rows:
                cells = [clean_text(c.text) for c in row.cells]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                order += 1
                blocks.append(Block(
                    text="[TABLO]\n" + "\n".join(rows),
                    page=None,
                    locator=f"Tablo {table_no}",
                    order=order,
                    extra={"type": "table"},
                ))

    if not blocks:
        raise RuntimeError(f"'{path.name}' içinde metin bulunamadı.")
    return blocks


def load_text(path: Path) -> List[Block]:
    enc = detect_encoding(path)
    try:
        raw = path.read_text(encoding=enc, errors="replace")
    except Exception:
        raw = path.read_text(encoding="utf-8", errors="replace")

    blocks: List[Block] = []
    # Boş satırla ayrılmış bölümleri ayrı bloklar olarak tut (konum takibi için)
    for i, section in enumerate(re.split(r"\n\s*\n", raw), start=1):
        text = clean_text(section)
        if text:
            blocks.append(Block(text=text, page=None, locator=f"Bölüm {i}", order=i))
    if not blocks:
        raise RuntimeError(f"'{path.name}' boş.")
    return blocks


def _table_blocks(rows: List[List[str]],
                  rows_per_chunk: int,
                  locator_fmt: str,
                  order_start: int = 0,
                  extra: Optional[Dict[str, str]] = None) -> List[Block]:
    """
    Tabloyu bloklara böler.

    KRİTİK TASARIM KARARI: rows_per_chunk=1 (varsayılan) ile HER SATIR AYRI BİR
    BLOK olur ve başlık satırı her bloğa kopyalanır. Neden?

    Tüm tabloyu tek parçaya koyarsanız, "2021'in 5. ayı" sorulduğunda model
    komşu satırı (6. ay) okuyup yanlış tutarı verebilir — klasik ve tehlikeli
    bir RAG hatasıdır. Satır bazlı parçalamada her satır bağımsız olarak
    vektörleşir; arama doğrudan doğru satırı bulur ve modele yalnızca o satır
    gösterilir. Karıştırılacak komşu satır bağlamda bulunmaz.

    Bedeli: parça sayısı artar (1000 satırlık tablo = 1000 parça). Çok büyük
    tablolarda config.yaml → retrieval.table_rows_per_chunk değerini
    5-10 yapabilirsiniz, ancak doğruluk bir miktar düşer.
    """
    if not rows:
        return []
    header = " | ".join(str(c).strip() for c in rows[0])
    blocks: List[Block] = []
    order = order_start
    step = max(1, int(rows_per_chunk))

    for start in range(1, len(rows), step):
        part = rows[start:start + step]
        body = "\n".join(" | ".join(str(c).strip() for c in r) for r in part)
        if not body.strip():
            continue
        order += 1
        end = start + len(part) - 1
        loc = locator_fmt.format(start=start, end=end)
        blocks.append(Block(
            text=f"[TABLO SATIRI] {header}\n{body}",
            page=None,
            locator=loc,
            order=order,
            extra={"type": "table", **(extra or {})},
        ))
    return blocks


def load_csv(path: Path, rows_per_chunk: int = 1) -> List[Block]:
    enc = detect_encoding(path)
    content = path.read_text(encoding=enc, errors="replace")
    try:
        dialect = csv.Sniffer().sniff(content[:4096], delimiters=",;\t|")
        delim = dialect.delimiter
    except Exception:
        delim = ";" if content[:4096].count(";") > content[:4096].count(",") else ","

    reader = csv.reader(io.StringIO(content), delimiter=delim)
    rows = [r for r in reader if any(str(c).strip() for c in r)]
    if not rows:
        raise RuntimeError(f"'{path.name}' boş.")

    fmt = "Satır {start}" if rows_per_chunk == 1 else "Satır {start}-{end}"
    return _table_blocks(rows, rows_per_chunk, fmt)


def _fmt_cell(v: object) -> str:
    """Hücreyi metne çevirir; sayı ve tarihleri Türkçe biçimde korur."""
    if v is None:
        return ""
    import datetime as _dt
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, float):
        # 1490.0 -> "1490", 1490.5 -> "1490,5"
        if v.is_integer():
            return str(int(v))
        return f"{v}".replace(".", ",")
    return str(v).strip()


def load_xls(path: Path, rows_per_chunk: int = 1) -> List[Block]:
    """Eski Excel biçimi (.xls). openpyxl bu biçimi okuyamaz, xlrd gerekir."""
    try:
        import xlrd  # type: ignore
    except ImportError:
        raise RuntimeError(
            f"'{path.name}' eski Excel biçimi (.xls). Okumak için 'xlrd' paketi gerekir: "
            "pip install xlrd   —  ya da dosyayı Excel'de açıp .xlsx olarak kaydedin."
        )

    book = xlrd.open_workbook(str(path))
    blocks: List[Block] = []
    order = 0
    for sheet in book.sheets():
        rows = []
        for r in range(sheet.nrows):
            vals = [_fmt_cell(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            if any(v for v in vals):
                rows.append(vals)
        if not rows:
            continue
        fmt = ("'" + sheet.name + "' sayfası, satır {start}" if rows_per_chunk == 1
               else "'" + sheet.name + "' sayfası, satır {start}-{end}")
        new = _table_blocks(rows, rows_per_chunk, fmt,
                            order_start=order, extra={"sheet": sheet.name})
        order += len(new)
        blocks.extend(new)

    if not blocks:
        raise RuntimeError(f"'{path.name}' içinde veri yok.")
    return blocks


def load_xlsx(path: Path, rows_per_chunk: int = 1) -> List[Block]:
    """Modern Excel (.xlsx / .xlsm / .xltx). Birleştirilmiş hücreler yayılır."""
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(str(path), read_only=False, data_only=True)
    blocks: List[Block] = []
    order = 0
    for ws in wb.worksheets:
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue  # gizli sayfalar genelde yardımcı/hesap sayfalarıdır

        grid = [[_fmt_cell(c) for c in row] for row in ws.iter_rows(values_only=True)]

        # Birleştirilmiş hücrelerde yalnızca sol-üst hücre dolu gelir;
        # başlıkların kaybolmaması için değeri tüm aralığa yayıyoruz.
        try:
            for rng in list(ws.merged_cells.ranges):
                r1, c1, r2, c2 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
                if r1 - 1 < len(grid) and c1 - 1 < len(grid[r1 - 1]):
                    val = grid[r1 - 1][c1 - 1]
                    if val:
                        for r in range(r1 - 1, min(r2, len(grid))):
                            for c in range(c1 - 1, min(c2, len(grid[r]))):
                                if not grid[r][c]:
                                    grid[r][c] = val
        except Exception:
            pass

        rows = [r for r in grid if any(r)]
        if not rows:
            continue
        fmt = ("'" + ws.title + "' sayfası, satır {start}" if rows_per_chunk == 1
               else "'" + ws.title + "' sayfası, satır {start}-{end}")
        new = _table_blocks(rows, rows_per_chunk, fmt,
                            order_start=order, extra={"sheet": ws.title})
        order += len(new)
        blocks.extend(new)
    wb.close()
    if not blocks:
        raise RuntimeError(f"'{path.name}' içinde veri yok.")
    return blocks


_DISPATCH = {
    ".pdf": load_pdf,
    ".docx": load_docx,
    ".txt": load_text,
    ".md": load_text,
    ".csv": load_csv,
    ".tsv": load_csv,
    ".xlsx": load_xlsx,
    ".xlsm": load_xlsx,
    ".xls": load_xls,
}

_TABLE_LOADERS = {".csv", ".tsv", ".xlsx", ".xlsm", ".xls"}


def load_document(path: Path,
                  table_rows_per_chunk: int = 1,
                  pdf_split_table_rows: bool = True,
                  ocr_options: Optional[Dict[str, object]] = None,
                  progress=None,
                  warnings: Optional[List[Dict[str, object]]] = None) -> List[Block]:
    ext = path.suffix.lower()
    if ext not in _DISPATCH:
        raise ValueError(f"Desteklenmeyen dosya türü: {ext}")
    if ext == ".pdf":
        return load_pdf(path, split_table_rows=pdf_split_table_rows,
                        ocr_options=ocr_options, progress=progress,
                        warnings=warnings)
    if ext in _TABLE_LOADERS:
        return _DISPATCH[ext](path, table_rows_per_chunk)  # type: ignore[call-arg]
    return _DISPATCH[ext](path)


def iter_documents(root: Path) -> Iterator[Path]:
    """Belge klasörünü özyinelemeli tarar, geçici/gizli dosyaları atlar."""
    for p in sorted(root.rglob("*")):
        if not p.is_file():
            continue
        if p.name.startswith((".", "~$")):
            continue
        if p.suffix.lower() in SUPPORTED_EXTENSIONS:
            yield p
